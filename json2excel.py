# -*- coding: utf-8 -*-
__author__ = 'faiyer'
import json
from openpyxl import *
import time

fp = open('../statics/aliyun.json', 'r')
data = json.loads(fp.read(), encoding='Unicode')
# print(len(data['items']))
# contestants = data['items']
rows = data['data']['main']['data']['disp_data']

outwb = Workbook()
outws = outwb.create_sheet('sheet')

# write header
keys = rows[0].keys()
for j, key in enumerate(keys):
    outws.cell(row=1, column=j+1).value = key

#write body
for i, row in enumerate(rows):
    for j, key in enumerate(keys):
        if key not in row:
          val = None
        else:
          val = row[key]

        if isinstance(val, list):
          val = ",".join(val)

        outws.cell(row=i+2, column=j+1).value = str(val)

now = time.time()
outwb.save(str(now) + 'aliyun.xlsx')

